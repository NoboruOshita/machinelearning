from django.http import HttpResponse, JsonResponse
from django.utils.translation import gettext as _
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.contrib import messages
from . models import *
from datetime import datetime
from openpyxl.styles import Alignment, Font
from django.db.models import Count, Case, When, Q
from django.utils.timezone import now, timedelta
from rest_framework import status
from dotenv import load_dotenv
import os
import requests
import joblib
import numpy as np
import openpyxl
import json
#load_dotenv()
from decouple import config

# VARIABLES GLOBALS
deviation_entropy = None
deviation_size = None
deviation_LBA = None
predictedProbabilityRF = None
predictedProbabilityXGB = None
selected_start = None
selected_end = None
#API_KEY = os.getenv('API_KEY')
API_KEY = config("API_KEY")

resultStorage = {
    'ransomware_type' : [],
    'reliability_percentage': [],
    'predictedProbabilityRF': [],
    'predictedProbabilityXGB': [],
    'deviation_entropy': [],
    'deviation_size': [],
    'deviation_LBA': []
}

def loginUser(request):
    '''
        - Login the User
    '''
    if request.method == 'POST':
        usuario = request.POST['username']
        psw = request.POST['password']
        user = authenticate(request, username = usuario, password = psw)
        if user is not None:
            login(request, user)
            messages.success(request, "Inicio de sesión exitoso. ¡Bienvenido!")
            return redirect('dashboard')
        else:
            messages.error(request, "Credenciales inválidas. Por favor, verifica tus credenciales e inténtalo nuevamente.")
            return redirect('login')
    return render(request, 'appWeb/login.html')

def logout_user(request):
    logout(request)
    messages.success(request, ("Sesion Cerrada Exitosamente"))
    return redirect('login')


def sign_up(request):
    '''
        - Create new User
        - Save the Data the new User in Database PostgreSQL
    '''
    if request.method == 'POST':
        usuario_nuevo = request.POST.get('username').strip()
        correo = request.POST.get('email').strip()
        pass_nuevo = request.POST.get('password').strip()
        pass_confirmacion = request.POST.get('confpassword').strip()

        #VALIDATION
        if not all([usuario_nuevo, correo, pass_nuevo, pass_confirmacion]):
            context = {'error_message': _('Todos los campos son obligatorios')}
            return render (request, 'appWeb/signup.html', context)
        if pass_nuevo != pass_confirmacion:
            context = {'error_message': _('Las contraseñas no coinciden')}
            return render (request, 'appWeb/signup.html', context)
        try:
            validate_email(correo)
        except ValidationError:
            context = {'error_message': _('El correo electrónico no es válido')}
            return render (request, 'appWeb/signup.html', context)
        if len(pass_nuevo) < 8:
            context = {'error_message': _('La contraseña debe tener al menos 8 caracteres')}
            print(f"Mensaje de error a mostrar: '{context['error_message']}'")
            return render(request, 'appWeb/signup.html', context)
        if auth_user.objects.filter(username=usuario_nuevo).exists() or auth_user.objects.filter(email=correo).exists():
            context = {'error_message': _('El nombre de usuario o el correo electrónico ya están en uso')}
            return render (request, 'appWeb/signup.html', context)
        
       # SAVE IN DATABASE
        try:
            user = User.objects.create_user(
                username=usuario_nuevo,
                email=correo,
                password=pass_nuevo
            )
            user.save()  # Saves the user in PostgreSQL
            messages.success(request, _('Se registró correctamente'))
            return redirect('login')
        except Exception as e:
            print('No se pudo registrar. Error: ', e)
            return JsonResponse({'Mensaje' : 'Error interno del servidor'}, status = 500)
    return render(request, 'appWeb/signup.html')

@login_required
def index(request):
    return render(request, 'appWeb/index.html')

@login_required
def detections(request):
    return render(request, 'appWeb/detection/detection.html')

@login_required
def dashboard(request):
    # -------------------------
    # Lógica de filtros para gráficas
    # -------------------------
    user = request.user
    endDate = now().date()
    startDate = endDate - timedelta(weeks=4)

    # Parámetros GET para gráfica de línea
    selected_start = request.GET.get('startDate')
    selected_end   = request.GET.get('endDate')
    if 'limpiar_filtro' in request.GET:
        startDate = endDate - timedelta(weeks=4)
    else:
        if selected_start:
            startDate = datetime.strptime(selected_start, '%Y-%m-%d').date()
        if selected_end:
            endDate = datetime.strptime(selected_end, '%Y-%m-%d').date()
        if startDate > endDate:
            endDate = startDate

    # Parámetro GET para gráfica de pastel
    start_date_pie_chart = request.GET.get('start_date_pie_chart')
    if 'limpiar_filtro_pie_chart' in request.GET or not start_date_pie_chart:
        start_date_pie_chart = endDate - timedelta(weeks=4)
    else:
        start_date_pie_chart = datetime.strptime(start_date_pie_chart, '%Y-%m-%d').date()

    base_responses = response.objects.filter(
        id_detection__id_data_ransomware__id_user=user.id,
        response_date__gte=start_date_pie_chart
    )
    # Datos pastel
    total_responses = base_responses.count()
    maligna_count   = base_responses.filter(action="Cuarentena").count()
    benigna_count   = base_responses.filter(action="Ransomware no detectado").count()
    maligna_pct     = (maligna_count / total_responses) * 100 if total_responses else 0
    benigna_pct     = (benigna_count / total_responses) * 100 if total_responses else 0
    pie_values      = json.dumps([round(maligna_pct, 2), round(benigna_pct, 2)])

    # Datos línea
    real_positive_qs = (
        response.objects
        .filter(id_detection__id_data_ransomware__id_user=user.id, response_date__range=(startDate, endDate))
        .values('response_date')
        .annotate(count=Count(Case(When(action="Cuarentena", then=1))))
        .order_by('response_date')
    )
    data = {
        'labels': [item['response_date'].strftime('%d/%m/%Y') for item in real_positive_qs],
        'counts': [item['count'] for item in real_positive_qs],
    }

    # -------------------------
    # Filtros para la tabla de eventos
    # -------------------------
    events_qs = detection.objects.filter(
        id_data_ransomware__id_user=user.id
    )

    table_start   = request.GET.get('table_start')
    table_end     = request.GET.get('table_end')
    result_filter = request.GET.get('result_filter')  # 'malicioso' / 'no_malicioso' / None

    # Filtrar por rango de fechas, solo si no es None ni cadena vacía
    if table_start and table_start not in ('None', ''):
        dt_start = datetime.strptime(table_start, '%Y-%m-%d').date()
        events_qs = events_qs.filter(detection_date__gte=dt_start)
    if table_end and table_end not in ('None', ''):
        dt_end = datetime.strptime(table_end, '%Y-%m-%d').date()
        events_qs = events_qs.filter(detection_date__lte=dt_end)

    # Filtrar por resultado
    if result_filter == 'malicioso':
        events_qs = events_qs.filter(
            Q(ransomware_detected=True) |
            Q(ransomware_type__icontains="Malicioso")
        )
    elif result_filter == 'no_malicioso':
        events_qs = events_qs.filter(
            ransomware_detected=False
        ).filter(
            Q(ransomware_type__isnull=True) | Q(ransomware_type='') | Q(ransomware_type='-')
        )

    # Paginación y orden
    events_qs   = events_qs.select_related('id_data_ransomware').order_by('-detection_date')
    paginator   = Paginator(events_qs, 12)
    page_number = request.GET.get('page')
    events_page = paginator.get_page(page_number)

    # -------------------------
    # Contexto y render
    # -------------------------
    context = {
        'data': data,
        'pie_values': pie_values,
        'startDate': startDate,
        'endDate': endDate,
        'start_date_pie_chart': start_date_pie_chart,
        'table_start': table_start,
        'table_end': table_end,
        'result_filter': result_filter,
        'events_page': events_page,
    }
    return render(request, 'appWeb/dashboard/dashboard.html', context)

@login_required
def predicRansomware(request):
    '''
        - Verifica el IP (Black List) almacenada en la BD
        - Si no hay IP en BD, continuar con el funcionamiento de ML
    '''
    rfModel = joblib.load('random_forest_model.pkl')
    xgbModel = joblib.load('xg_boost_model.pkl')
    label_encoder = joblib.load('label_encoder.pkl')
    
    id_user = request.user.id
    user_instance = auth_user.objects.get(id=id_user)
    id_blacklist_id = None

    context = {
        'Timestamp_s': '',
        'Timestamp_ms': '',
        'LBA': '',
        'Size': '',
        'Entropy': '',
        'IP': ''
    }

    if request.method == 'POST':
        ts_raw  = request.POST.get('Timestamp_s', '').strip()
        tms_raw = request.POST.get('Timestamp_ms', '').strip()
        lba_raw = request.POST.get('LBA', '').strip()
        sz_raw  = request.POST.get('Size', '').strip()
        ent_raw = request.POST.get('Entropy', '').strip()
        ip_raw  = request.POST.get('IP', '').strip()
        
        # volcamos al contexto para que se vuelvan a mostrar si hay error
        context.update({
            'Timestamp_s': ts_raw,
            'Timestamp_ms': tms_raw,
            'LBA': lba_raw,
            'Size': sz_raw,
            'Entropy': ent_raw,
            'IP': ip_raw,
        })
        print(context)
        # 1) Validar que los campos obligatorios no estén vacíos
        if not (ts_raw and tms_raw and lba_raw and sz_raw and ent_raw):
            context['error_message'] = _("Tienes que completar todos los campos.")
            return render(request, 'appWeb/detection/detection.html', context)

        # 2) Validar longitud exacta de timestamps
        if len(ts_raw) < 10:
            print(ts_raw)
            context['error_message'] = _("El campo de Registro de Tiempo debe ser numérico y contener exactamente 10 dígitos (segundos)")
            return render(request, 'appWeb/detection/detection.html', context)

        if len(tms_raw) < 9:
            context['error_message'] = _("El campo de Registro de Tiempo debe ser numérico y contener al menos 9 dígitos (milisegundos)")
            return render(request, 'appWeb/detection/detection.html', context)

        if len(lba_raw) < 7:
            context['error_message'] = _("El campo Dirección de Bloqueo Lógico debe ser numérico y contener al menos 7 dígitos")
            return render(request, 'appWeb/detection/detection.html', context)

        if len(sz_raw) < 4:
            context['error_message'] = _("El campo Tamaño del Bloqueo deber ser numérico y contener exactamente 4 dígitos")
            return render(request, 'appWeb/detection/detection.html', context)

        if len(ent_raw) < 16:
            context['error_message'] = _("El campo Entropía de Shannon debe ser numérico decimal y contener al menos 16 dígitos")
            return render(request, 'appWeb/detection/detection.html', context)

        # 3) Intentar convertir a tipos numéricos
        try:
            timestampS  = int(ts_raw)
            timestampMS = int(tms_raw)
            LBA         = float(lba_raw)
            size        = float(sz_raw)
            entropy     = float(ent_raw)
            IP          = ip_raw or ''
        except ValueError:
            context['error_message'] = _("Por favor, ingrese valores válidos en los campos.")
            return render(request, 'appWeb/detection/detection.html', context)

        
        predictedFinal = 0
        date = datetime.today().date().strftime('%m/%d/%Y')

        # Verificar la ip de BD (Blacklist of BD us)
        if IP != '' and blacklist.objects.filter(IP=IP).exists():
            blacklist_instance = blacklist.objects.get(IP=IP)
            ransomwareType = blacklist_instance.ransomware_type
            id_blacklist = blacklist_instance.id_blacklist
            predictedFinal = 100
            context = {'ransomware_type': ransomwareType,
                'probability': predictedFinal,
                'date': date,
                'motive':  [_("La ip se encuentra dentro del backlist")]
            }

            # Storage in DB
            saveDataRansomware(
                id_user_id=user_instance, 
                id_blacklist_id=blacklist_instance,
                timestamp_s=timestampS, 
                timestamp_ms=timestampMS, 
                lba=LBA, 
                block_size=size, 
                entropy_shannon=entropy
            )            
            return render(request,'appWeb/detection/detection.html',context)
        
        # Verificar la ip por servicio AbuseIPDB (Use service api)
        if IP != '':
            try:
                url = "https://api.abuseipdb.com/api/v2/check"
                api_key = API_KEY

                information = {
                    "ipAddress": str(IP),
                    "maxAgeInDays": "90"
                }

                api = {
                    "key": api_key,
                    "Accept": "application/json"
                }

                response = requests.get(url, headers=api, params=information)
                res = response.json()
                probability = res['data']['abuseConfidenceScore']
                
                if probability >= 70:
                    context = {
                        'ransomware_type': ' Ip maliciosa ransomware',
                        'probability': probability,
                        'date': date,
                        'motive': [_("Ip maliciosa de ransomware detectada por AbuseIPDB")]
                    }
                    saveDetection(None, 'Malicioso por AbuseIPDB', probability, date)
                    return render(request,'appWeb/detection/detection.html',context)
            except Exception as e:
                print("Error", e)
                return JsonResponse({'mensaje': 'Error interno del servidor'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Verificar que todos los campos estén llenos
        if not (timestampS and timestampMS and LBA and size and entropy):
            context = {"error_message": _("Tienes que completar todos los campos.")}
            return render(request, 'appWeb/detection/detection.html', context)

        # try:
        #     # Convertir los campos a los tipos adecuados
        #     timestampS = int(timestampS)
        #     timestampMS = int(timestampMS)
        #     LBA = float(LBA)
        #     size = float(size)
        #     entropy = float(entropy)
        # except ValueError:
        #     context = {"error_message": _("Por favor, ingrese valores válidos en los campos")}
        #     return render(request, 'appWeb/detection/detection.html', context)

        # Verify data with the ML Model
        inputData = np.array([[timestampS, timestampMS, LBA, size, entropy]])
        
        '''
        Probabilidad estimada por cada modelo (Random Forest y XGBoost):
        Explicación:
            Se calcula la probabilidad final combinada de ambos modelos. Se muestrar estas probabilidades por separado como un valor cuantitativo que respalda la decisión. 
        Ejemplo:
            "Random Forest estima una probabilidad de detección del X%, mientras que XGBoost estima una probabilidad del Y%".
        '''
        global predictedProbabilityRF, predictedProbabilityXGB
        # Prediction Model Random Forest
        probabilitiesRF = rfModel.predict_proba(inputData)
        predictedProbabilityRF = np.max(probabilitiesRF) * 100
        typeRansomwareRF = rfModel.predict(inputData)
        
        # Prediction Model Extreme Gradient Boosting
        probabilitiesXGB = xgbModel.predict_proba(inputData)
        predictedProbabilityXGB = np.max(probabilitiesXGB) * 100
        preTypeRansomwareXGB = xgbModel.predict(inputData)
        typeRansomwareXGB = label_encoder.inverse_transform(preTypeRansomwareXGB)

        '''
        Desviación o anomalía en las características:
        Explicación
            Para cada entrada proporcionada, se calcula la desviación respecto a un valor promedio de las características normales. Así, se justifica el porcentaje en base a cuán lejos están los valores del comportamiento esperado (valores típicos de ransomware vs normales).
        Ejemplo:
            "La entropía observada es un Z% mayor que el promedio de los archivos normales, lo que contribuye a la alta probabilidad de ransomware".
        '''
        # Analyze deviations (anomalies)
        global deviation_entropy, deviation_size, deviation_LBA
        deviation_entropy = round(abs(entropy - 0.95) * 100, 2)  # 0.95 as average entropy threshold
        deviation_size = round(abs(size - 1024) / 100, 2) # 1024 as a threshold for block size
        deviation_LBA = round(abs(LBA - 600000), 2)  # LBA threshold

        # Final prediction
        if typeRansomwareRF[0] == typeRansomwareXGB[0]:
            ransomwareType = typeRansomwareXGB[0]
            # Promedio Simple
            predictedFinal = round((predictedProbabilityRF + predictedProbabilityXGB) / 2, 2)
        elif predictedProbabilityRF > predictedProbabilityXGB:
            ransomwareType = typeRansomwareRF[0]
            predictedFinal = round(predictedProbabilityRF, 2)
        else:
            ransomwareType = typeRansomwareXGB[0]
            predictedFinal = round(predictedProbabilityXGB, 2)

        # Evaluate final confidence and response
        data_ransomware_instance = None
        detection_instance = None
        response_instance = None

        features = {'entropy': entropy, 'size': size, 'LBA': LBA}
        mensaje, motive = responseRansomware(predictedFinal, features)
        
        # Storage in DB
        data_ransomware_instance =saveDataRansomware(id_user_id=user_instance, id_blacklist_id=None, timestamp_s=timestampS, timestamp_ms=timestampMS, lba=LBA, block_size=size, entropy_shannon=entropy)
        detection_instance = saveDetection(int(data_ransomware_instance), ransomwareType, predictedFinal, date)

        if predictedFinal < 70:  motive = None
        response_instance = saveResponse (int(detection_instance), motive,date)
        log(int(detection_instance),int(response_instance))
        
        ransomwareType = ransomwareType if predictedFinal > 70 else ""
        global resultStorage
        resultStorage['ransomware_type'].append(ransomwareType)
        resultStorage['reliability_percentage'].append(predictedFinal)
        resultStorage['predictedProbabilityRF'].append(predictedProbabilityRF)
        resultStorage['predictedProbabilityXGB'].append(predictedProbabilityXGB)
        resultStorage['deviation_entropy'].append(deviation_entropy)
        resultStorage['deviation_size'].append(deviation_size)
        resultStorage['deviation_LBA'].append(deviation_LBA)

        if predictedFinal < 70: motive = _('Sin respuesta contra la amenaza')
        context.update({
            'ransomware_type': ransomwareType,
            'probability': predictedFinal,
            'date': date,
            'motive': motive,
            'predictedProbabilityRF': round(predictedProbabilityRF,2),
            'predictedProbabilityXGB': round(predictedProbabilityXGB,2),
            'deviation_entropy': deviation_entropy,
            'deviation_size': deviation_size,
            'deviation_LBA': deviation_LBA
        })

    return render(request, 'appWeb/detection/detection.html', context)

def responseRansomware(predictedFinal, features):
    '''
        - Analisis de umbrales de cada caracteristicas (Anomalia)
    '''
    motivos = []
    if features['entropy'] > 0.95:
        motivos.append(_("Se detectó anomalía en la entropía de Shannon"))
    
    if features['size'] > 1024:
        motivos.append(_("El tamaño del bloque es inusualmente grande"))
    
    if features['LBA'] > 500000:
        motivos.append(_("Acceso a bloques sospechosos de memoria"))
    
    if not motivos:
        motivos.append(_("No se detectaron anomalías significativas en las características observadas"))

    mensaje = f"El documento fue puesto en cuarentena. Motivos:\n"
    for motivo in motivos:
        mensaje += f"- {motivo}\n"
    
    return mensaje, motivos

# Storages in DataBase
def saveDataRansomware(id_user_id, id_blacklist_id, timestamp_s, timestamp_ms, lba, block_size, entropy_shannon):
    try:
        data = data_ransomware(
            id_user=id_user_id,
            id_blacklist=id_blacklist_id,
            timestamp_s=timestamp_s,
            timestamp_ms=timestamp_ms,
            lba=lba,
            block_size=block_size,
            entropy_shannon=entropy_shannon
        )
        
        data.save()
        return data.id_data_ransomware
    except Exception as e:
        print("Error al guardar en data_ransomware:", e)
        return None

def saveDetection(data_ransomware_instance, ransomwareType, predictedFinal, date):
    ransomware_detected = data_ransomware_instance is not None
    if predictedFinal < 70:
        ransomwareType = None
        ransomware_detected = False
    date = datetime.strptime(date, '%m/%d/%Y').date()

    data_ransomware_obj = None
    if data_ransomware_instance is not None:
        data_ransomware_obj = data_ransomware.objects.get(id_data_ransomware=data_ransomware_instance)

    data = detection(
        id_data_ransomware=data_ransomware_obj,
        ransomware_detected=ransomware_detected,
        ransomware_type=ransomwareType,
        percentage_reliability=predictedFinal,
        detection_date=date
    )
    data.save()

    return data.id_detection

def saveResponse(detection_instance, motive, date):
    date = datetime.strptime(date, '%m/%d/%Y').date()
    action = "Cuarentena" if motive != None else None
    if action == None: 
        detail = 'Ninguna'
        action = 'Ransomware no detectado'
    else:
        detail = ','.join(motive).replace(',', '. ')
    data = response(
        id_detection=detection.objects.get(id_detection=detection_instance),
        action=action,
        detail=detail,
        response_date= date
    )
    data.save()
    
    return data.id_response

def log(detection_instance, response_instance):
    log = logs(
        id_detection=detection.objects.get(id_detection=detection_instance),
        id_response=response.objects.get(id_response=response_instance)
    )
    log.save()

@login_required
def excelDetail(request):
    date = datetime.today()
    formatDate = datetime.strftime(date, '%d/%m/%y')
    formatHours = datetime.now().strftime("%H:%M:%S")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    entityName = {
        'Deviation Entropy': _("Desviación de Entropía"),
        'Deviation size': _("Desviación de Tamaño"),
        'Deviation LBA': _("Desviacion de Dirección de bloque lógico")
    }
    global resultStorage, deviation_size, deviation_LBA, deviation_entropy

    # Column headings
    headers = [
        _("Fecha"),
        _("Hora"),
        _("Ransomware"),
        _("Porcentaje de confiabilidad"),
        _("Probabilidad de predicción de Random Forest"), 
        _("Probalidad de predicción de XGBoost"), 
        entityName['Deviation size'], 
        entityName['Deviation LBA'], 
        entityName['Deviation Entropy'],
    ]

    # Add headers to Excel 
    sheet.append(headers)
    
    # Add the date to each row of data 
    rows = [] 
    for i in range(len(resultStorage['predictedProbabilityRF'])): 
        # Add the data for each “input” to a row 
        row = [ formatDate,
               formatHours,
               resultStorage['ransomware_type'][i],
               resultStorage['reliability_percentage'][i],
               resultStorage['predictedProbabilityRF'][i], 
               resultStorage['predictedProbabilityXGB'][i],
               resultStorage['deviation_size'][i], 
               resultStorage['deviation_LBA'][i], 
               resultStorage['deviation_entropy'][i] ] 
        # Add the row to the row list 
        rows. append(row) 
    
    # Add all rows to Excel.
    for row in rows:
        sheet.append(row)
    

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = _("Detalle del resultado") + ".xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)

    print("¡Archivo Excel creado exitosamente!")

    #Clear resultStorage
    for key in resultStorage:
        resultStorage[key].clear()
        
    return response

@login_required
def getExcelflowChart(request):
    global selected_start, selected_end

    # Check and convert selected dates
    if selected_start and selected_end:
        start_date = datetime.strptime(selected_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(selected_end, '%Y-%m-%d').date()
    else:
        end_date = datetime.now().date()
        start_date = end_date - timedelta(weeks=4)
    
    # Query the data in the date range
    real_positive_data = (
        response.objects
        .filter(response_date__range=(start_date, end_date))
        .values('response_date', 'action', 'detail')
        .annotate(count=Count('id_response'))
        .order_by('response_date')
    )

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("Reporte de Reales Positivos")

    # Set headers
    headers = ["Fecha", "Cantidad de Ransomware Detectados", "Acción Tomada", "Detalles"]
    ws.append(headers)

    # Formatting headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Filling in the data in the Excel file
    for record in real_positive_data:
        ws.append([
            record['response_date'].strftime('%d/%m/%Y'),
            record['count'],  # Number of ransomware detected
            record['action'], # Action taken
            record['detail']   # Details
        ])

    # Configurar la respuesta HTTP para la descarga de archivo
    responses = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    responses['Content-Disposition'] = f'attachment; filename="reporte_reales_positivos_{start_date}_a_{end_date}.xlsx"'

    # Save the file in the HTTP response
    wb.save(responses)

    return responses